"""
giris_exporter.py
=================
Fills the user's "Giriş / Çıktı" Excel template with data from the DB.

What we write (never touching formula cells):
  Giriş sheet:
    H2          → dönem_adı          (period name, e.g. "ŞUBAT 2026")
    H26         → rate_per_meter      (USD/m)
    H53         → exchange_rate       (USD/TL)
    H41         → kuyuda_kalan        (USD, material left in hole)
    C33:C36     → target metres per machine (GEO 900E-1/2/3/5)
    G45:G48     → standby unit price  (USD/hr, all machines same)
    A5:F24      → borehole rows (up to 20):
                    A=rig, B=hole, C=start_date, D=end_date,
                    E=m_start, F=m_end
                  Rows beyond actual data are cleared (A:F set to "")
                  G and H columns are left untouched (they hold formulas)

  "Bekleme Süresi Aciklama" sheet (created if missing):
    H2 → net standby hours GEO 900E-1
    H3 → net standby hours GEO 900E-2
    H4 → net standby hours GEO 900E-3
    H5 → net standby hours GEO 900E-5
    H6 → =SUM(H2:H5)   (written as a formula)

Formula cells (G5:G24, H5:H24, H25, H27-H29, H33-H37, H42, H45-H49,
H54, H56-H57, Çıktı sheet) are never touched — Excel recalculates them
when the file is opened.
"""

from pathlib import Path
from openpyxl import load_workbook

import db.models as m

MACHINES = ["GEO 900E-1", "GEO 900E-2", "GEO 900E-3", "GEO 900E-5"]
GIRIS_SHEET   = "Giriş"
BEKLEME_SHEET = "Bekleme Süresi Aciklama"
MAX_BOREHOLE_ROWS = 20   # rows 5..24 in the template


def fill_template(template_path: str, output_path: str,
                  contractor_id: int, month: int, year: int) -> None:
    """
    Load *template_path*, write all input values, save to *output_path*.
    Does NOT modify *template_path*.

    Raises FileNotFoundError if template_path does not exist.
    Raises ValueError if the "Giriş" sheet is not found in the template.
    """
    tp = Path(template_path)
    if not tp.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = load_workbook(tp)

    if GIRIS_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{GIRIS_SHEET}' not found in template. "
            f"Available sheets: {wb.sheetnames}"
        )

    # ── Fetch data ──────────────────────────────────────────────────────────
    contractor   = m.get_contractor(contractor_id)
    if not contractor:
        raise ValueError(f"Contractor {contractor_id} not found.")

    ps           = m.get_period_settings(contractor_id, month, year)
    bh_entries   = m.get_drilling_entries(contractor_id, month, year)
    sb_net_hours = m.get_standby_net_hours_per_rig(contractor_id, month, year)

    rate_per_meter = contractor["rate_per_meter"]

    # ── Write to Giriş sheet ───────────────────────────────────────────────
    ws = wb[GIRIS_SHEET]

    ws["H2"]  = ps.get("donem_adi", "")
    ws["H26"] = rate_per_meter
    ws["H41"] = ps.get("kuyuda_kalan", 0.0)
    ws["H53"] = ps.get("exchange_rate", 0.0)

    # Target metres per machine (Tablo-2)
    target_map = {
        "GEO 900E-1": ("C33", ps.get("target_geo1", 700.0)),
        "GEO 900E-2": ("C34", ps.get("target_geo2", 700.0)),
        "GEO 900E-3": ("C35", ps.get("target_geo3", 700.0)),
        "GEO 900E-5": ("C36", ps.get("target_geo5", 700.0)),
    }
    for machine, (cell, val) in target_map.items():
        ws[cell] = val

    # Standby unit price (Tablo-4, all machines same rate)
    sb_rate = ps.get("standby_rate", 75.0)
    for cell in ("G45", "G46", "G47", "G48"):
        ws[cell] = sb_rate

    # Borehole rows (Tablo-1): rows 5..24 → columns A..F
    for i, entry in enumerate(bh_entries[:MAX_BOREHOLE_ROWS]):
        r = 5 + i
        ws.cell(r, 1).value = entry["rig_name"]   or ""
        ws.cell(r, 2).value = entry["hole_id"]    or ""
        ws.cell(r, 3).value = entry["start_date"] or ""
        ws.cell(r, 4).value = entry["end_date"]   or ""
        ws.cell(r, 5).value = entry["start_depth"]
        ws.cell(r, 6).value = entry["end_depth"]

    # Clear any extra rows (buffer)
    for i in range(len(bh_entries), MAX_BOREHOLE_ROWS):
        r = 5 + i
        for col in range(1, 7):   # A..F only
            ws.cell(r, col).value = ""

    # ── Bekleme Süresi Aciklama sheet ──────────────────────────────────────
    if BEKLEME_SHEET in wb.sheetnames:
        ws_bk = wb[BEKLEME_SHEET]
    else:
        ws_bk = wb.create_sheet(title=BEKLEME_SHEET)

    ws_bk["H2"] = sb_net_hours.get("GEO 900E-1", 0.0)
    ws_bk["H3"] = sb_net_hours.get("GEO 900E-2", 0.0)
    ws_bk["H4"] = sb_net_hours.get("GEO 900E-3", 0.0)
    ws_bk["H5"] = sb_net_hours.get("GEO 900E-5", 0.0)
    ws_bk["H6"] = "=SUM(H2:H5)"

    # ── Save ───────────────────────────────────────────────────────────────
    wb.save(output_path)
