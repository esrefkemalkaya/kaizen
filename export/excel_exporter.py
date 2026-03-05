"""
Generates a monthly drilling invoice Excel workbook.

Layout:
  - Summary sheet (index 0, when >1 contractor): totals per contractor
  - Per contractor:
      - Main sheet   : header + Tablo-1 drilling detail + standby/PPE subtotals only
      - "Name - SB"  : full Standby Hours detail table
      - "Name - PPE" : full PPE / Diesel detail table
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

import db.models as m

MONTHS = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]

# ── Color palette ──────────────────────────────────────────────────────────────
C_HEADER_BG   = "1E2A38"
C_SECTION_BG  = "37474F"
C_COL_HDR_BG  = "1565C0"
C_STANDBY_BG  = "E65100"
C_DEDUCT_BG   = "B71C1C"
C_SUBTOTAL_BG = "E3F2FD"
C_NET_BG      = "E8F5E9"
C_WHITE       = "FFFFFF"
C_BLACK       = "000000"

_thin  = Side(style="thin",   color=C_BLACK)
_thick = Side(style="medium", color=C_BLACK)

# Column layout (9 cols):
# 1:#  2:ID/Name  3:StartDate/Date  4:EndDate/Start  5:SDepth/End  6:EDepth/Type  7:Meters/Hours  8:Rate  9:Amount
NCOLS = 9
COL_WIDTHS = [5, 18, 11, 11, 11, 22, 10, 13, 16]


def thin_border():
    return Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def thick_bottom():
    return Border(left=_thin, right=_thin, top=_thin, bottom=_thick)


def _cell(ws, row, col, value="", bold=False, italic=False,
          fg=C_BLACK, bg=None, align="left", fmt=None, size=11):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, italic=italic, color=fg, size=size)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = thin_border()
    if fmt:
        cell.number_format = fmt
    return cell


def _section_header(ws, row, label, bg=C_SECTION_BG):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(bold=True, color=C_WHITE, size=12)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thin_border()
    return row + 1


def _col_headers(ws, row, labels, bg):
    for col, h in enumerate(labels, 1):
        _cell(ws, row, col, h, bold=True, fg=C_WHITE, bg=bg, align="center")
    return row + 1


def _sheet_name(base, suffix, max_len=31):
    """Truncate base so that 'base - suffix' fits within max_len."""
    full = f"{base} - {suffix}"
    if len(full) <= max_len:
        return full
    trim = max_len - len(suffix) - 3  # 3 = " - "
    return f"{base[:trim]} - {suffix}"


def generate_invoice(project_id, contractor_id, month, year, output_path,
                     usd_tl_rate: float = 1.0):
    project_rows = [p for p in m.get_projects() if p["id"] == project_id]
    if not project_rows:
        raise ValueError("Project not found.")
    project = project_rows[0]

    wb = Workbook()
    wb.remove(wb.active)

    contractors = m.get_contractors(project_id)
    summary_data = []

    for contractor in contractors:
        if contractor_id != -1 and contractor["id"] != contractor_id:
            continue
        _build_contractor_sheets(wb, project, contractor, month, year, usd_tl_rate, summary_data)

    if len(summary_data) > 1:
        _build_summary_sheet(wb, summary_data, project, month, year, usd_tl_rate)

    if not wb.worksheets:
        raise ValueError("No data to export.")

    wb.save(output_path)


# ── Per-contractor sheet builder ───────────────────────────────────────────────

def _build_contractor_sheets(wb, project, contractor, month, year, usd_tl_rate, summary_data):
    cname  = contractor["name"]
    ctype  = contractor["type"]
    rate_m = contractor["rate_per_meter"]
    rate_s = contractor["standby_hour_rate"]

    borehole_entries = m.get_drilling_entries(contractor["id"], month, year)
    standby_entries  = m.get_standby_entries(contractor["id"], month, year)
    ppe_rows = m.get_ppe_charges(contractor["id"], month, year)   if ctype == "underground" else []
    dsl_rows = m.get_diesel_charges(contractor["id"], month, year) if ctype == "surface"     else []
    charge_rows = ppe_rows if ctype == "underground" else dsl_rows

    # ── Calculate totals ───────────────────────────────────────────────────────
    total_meters = sum(e["meters_drilled"] for e in borehole_entries)
    total_borehole = total_meters * rate_m

    # Apply blasting deduction rule: each rig gets 24 free Patlatma hrs/month
    rig_summary, total_payable_sb_hours, total_standby = \
        m.calc_standby_rig_summary(standby_entries, rate_s)
    total_sb_hours_raw = sum(e["hours"] for e in standby_entries)

    total_deductions_tl  = sum(r["quantity"] * r["unit_price"] for r in charge_rows)
    total_deductions_usd = total_deductions_tl / usd_tl_rate if usd_tl_rate else 0.0

    total_work = total_borehole + total_standby
    net_usd    = total_work - total_deductions_usd
    net_tl     = net_usd * usd_tl_rate

    # ── Build the three sheets ─────────────────────────────────────────────────
    ws_main = wb.create_sheet(title=cname[:31])
    _build_main_sheet(ws_main, project, contractor, month, year, usd_tl_rate,
                      borehole_entries, rate_m,
                      total_meters, total_borehole,
                      rig_summary, total_payable_sb_hours, total_standby,
                      total_deductions_tl, total_deductions_usd,
                      total_work, net_usd, net_tl, ctype)

    ws_sb = wb.create_sheet(title=_sheet_name(cname[:20], "Standby"))
    _build_standby_sheet(ws_sb, cname, month, year, standby_entries, rate_s,
                         rig_summary, total_payable_sb_hours, total_standby)

    deduct_label = "PPE" if ctype == "underground" else "Diesel"
    ws_ppe = wb.create_sheet(title=_sheet_name(cname[:20], deduct_label))
    _build_deductions_sheet(ws_ppe, cname, ctype, month, year, charge_rows,
                            total_deductions_tl, total_deductions_usd, usd_tl_rate)

    summary_data.append({
        "name":       cname,
        "type":       ctype,
        "boreholes":  total_borehole,
        "standby":    total_standby,
        "deduct_tl":  total_deductions_tl,
        "deduct_usd": total_deductions_usd,
        "net_usd":    net_usd,
        "net_tl":     net_tl,
    })


# ── Main contractor sheet (drilling detail + subtotals only) ───────────────────

def _build_main_sheet(ws, project, contractor, month, year, usd_tl_rate,
                      borehole_entries, rate_m,
                      total_meters, total_borehole,
                      rig_summary, total_payable_sb_hours, total_standby,
                      total_deductions_tl, total_deductions_usd,
                      total_work, net_usd, net_tl, ctype):
    cname  = contractor["name"]
    rate_s = contractor["standby_hour_rate"]

    ws.sheet_view.showGridLines = False
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── Invoice header ─────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 40
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row=row, column=1, value="DRILLING INVOICE")
    c.font = Font(bold=True, size=18, color=C_WHITE)
    c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    for label, value in [
        ("Project",       project["name"]),
        ("Location",      project["location"] or ""),
        ("Contractor",    cname),
        ("Type",          ctype.capitalize()),
        ("Period",        f"{MONTHS[month - 1]} {year}"),
        ("Exchange Rate", f"1 USD = {usd_tl_rate:,.4f} TL"),
        ("Generated",     date.today().strftime("%d %b %Y")),
    ]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        _cell(ws, row, 1, label, bold=True, bg="ECEFF1", align="right")
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=NCOLS)
        _cell(ws, row, 3, value)
        row += 1

    row += 1  # spacer

    # ── Tablo-1: BOREHOLE DRILLING (full detail) ───────────────────────────────
    row = _section_header(ws, row, "  TABLO-1 — BOREHOLE DRILLING")
    row = _col_headers(ws, row, [
        "#", "Hole ID", "Start Date", "End Date",
        "Start Depth (m)", "End Depth (m)", "Meters", "Rate / m (USD)", "Amount (USD)"
    ], C_COL_HDR_BG)

    for idx, e in enumerate(borehole_entries, 1):
        meters = e["meters_drilled"]
        amount = meters * rate_m
        _cell(ws, row, 1, idx, align="center")
        _cell(ws, row, 2, e["hole_id"])
        _cell(ws, row, 3, e["start_date"],  align="center")
        _cell(ws, row, 4, e["end_date"],    align="center")
        _cell(ws, row, 5, e["start_depth"], align="right", fmt="#,##0.00")
        _cell(ws, row, 6, e["end_depth"],   align="right", fmt="#,##0.00")
        _cell(ws, row, 7, meters, align="right", fmt="#,##0.00")
        _cell(ws, row, 8, rate_m, align="right", fmt='"$"#,##0.00')
        _cell(ws, row, 9, amount, align="right", fmt='"$"#,##0.00', bold=True)
        row += 1

    if not borehole_entries:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        _cell(ws, row, 1, "No borehole entries for this period.", italic=True, align="center")
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    _cell(ws, row, 1, "BOREHOLE SUBTOTAL", bold=True, bg=C_SUBTOTAL_BG, align="right")
    _cell(ws, row, 7, total_meters,   align="right", fmt="#,##0.00",     bold=True, bg=C_SUBTOTAL_BG)
    _cell(ws, row, 8, "",             bg=C_SUBTOTAL_BG)
    _cell(ws, row, 9, total_borehole, align="right", fmt='"$"#,##0.00', bold=True, bg=C_SUBTOTAL_BG)
    row += 2

    # ── Tablo-2: STANDBY HOURS — per-rig summary, details on separate sheet ──────
    row = _section_header(ws, row,
                          "  TABLO-2 — STANDBY HOURS  (see 'Standby' sheet for full detail)",
                          bg=C_STANDBY_BG)
    # Per-rig columns: Rig | Non-Blast (h) | Blast Total (h) | −24h Free | Blast Paid (h) | Net Payable (h) | Rate | Amount
    row = _col_headers(ws, row, [
        "Drill Rig", "Non-Blasting (h)", "Patlatma Total (h)",
        "−24h Free", "Patlatma Paid (h)", "Net Payable (h)", "Rate / hr", "Amount (USD)", ""
    ], C_STANDBY_BG)

    if not rig_summary:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        _cell(ws, row, 1, "No standby entries for this period.", italic=True, align="center")
        row += 1
    else:
        for r in rig_summary:
            _cell(ws, row, 1, r["rig"], bold=True, bg="FFF3E0")
            _cell(ws, row, 2, r["other"],          align="right", fmt="#,##0.00", bg="FFF3E0")
            _cell(ws, row, 3, r["blasting_total"],  align="right", fmt="#,##0.00", bg="FFF3E0")
            _cell(ws, row, 4, -r["blasting_free"],  align="right", fmt="#,##0.00", bg="FFF3E0",
                  fg="B71C1C")
            _cell(ws, row, 5, r["blasting_paid"],   align="right", fmt="#,##0.00", bg="FFF3E0")
            _cell(ws, row, 6, r["payable_hours"],   align="right", fmt="#,##0.00", bg="FFF3E0",
                  bold=True)
            _cell(ws, row, 7, rate_s, align="right", fmt='"$"#,##0.00', bg="FFF3E0")
            _cell(ws, row, 8, r["amount"], align="right", fmt='"$"#,##0.00', bg="FFF3E0",
                  bold=True)
            _cell(ws, row, 9, "", bg="FFF3E0")
            row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    _cell(ws, row, 1, "STANDBY SUBTOTAL (NET PAYABLE)", bold=True, bg="FFE0B2",
          align="right", fg=C_STANDBY_BG)
    _cell(ws, row, 6, total_payable_sb_hours, align="right", fmt="#,##0.00",
          bold=True, bg="FFE0B2")
    _cell(ws, row, 7, "",          bg="FFE0B2")
    _cell(ws, row, 8, total_standby, align="right", fmt='"$"#,##0.00', bold=True,
          bg="FFE0B2", fg=C_STANDBY_BG)
    _cell(ws, row, 9, "",          bg="FFE0B2")
    row += 2

    # ── Tablo-3: DEDUCTIONS — totals only, detail on separate sheet ────────────
    deduct_label = "PPE CHARGES (DEDUCTION — TL)" if ctype == "underground" else "DIESEL CHARGES (DEDUCTION — TL)"
    deduct_sheet = "PPE" if ctype == "underground" else "Diesel"
    row = _section_header(ws, row,
                          f"  TABLO-3 — {deduct_label}  (see '{deduct_sheet}' sheet for details)",
                          bg=C_DEDUCT_BG)
    row = _col_headers(ws, row, [
        "", "", "", "", "", "", "", "Total (TL)", "Total (USD)"
    ], C_DEDUCT_BG)

    if not total_deductions_tl:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        _cell(ws, row, 1, "No charges for this period.", italic=True, align="center")
        row += 1
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        _cell(ws, row, 1, f"See '{deduct_sheet}' sheet for full breakdown", italic=True,
              align="center", bg="FFEBEE")
        _cell(ws, row, 8, total_deductions_tl,  align="right", fmt='"₺"#,##0.00', bg="FFEBEE")
        _cell(ws, row, 9, total_deductions_usd, align="right", fmt='"$"#,##0.00', bold=True,
              bg="FFEBEE")
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    _cell(ws, row, 1, "DEDUCTION SUBTOTAL", bold=True, bg="FFCDD2", align="right")
    _cell(ws, row, 8, total_deductions_tl,
          align="right", fmt='"₺"#,##0.00', bold=True, bg="FFCDD2", fg=C_DEDUCT_BG)
    _cell(ws, row, 9, total_deductions_usd,
          align="right", fmt='"$"#,##0.00', bold=True, bg="FFCDD2", fg=C_DEDUCT_BG)
    row += 2

    # ── Net Payable ────────────────────────────────────────────────────────────
    def _net_row(label, value, fmt, bg, fg, size=12, height=26):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(bold=True, size=size, color=C_WHITE)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = thick_bottom()
        nc = ws.cell(row=row, column=9, value=value)
        nc.font = Font(bold=True, size=size, color=fg)
        nc.fill = PatternFill("solid", fgColor=C_NET_BG)
        nc.alignment = Alignment(horizontal="right", vertical="center")
        nc.border = thick_bottom()
        nc.number_format = fmt
        ws.row_dimensions[row].height = height
        row += 1

    _net_row("TOTAL WORK (USD)",       total_work,          '"$"#,##0.00', "455A64", "37474F")
    _net_row("TOTAL DEDUCTIONS (TL)",  total_deductions_tl, '"₺"#,##0.00', C_DEDUCT_BG, C_DEDUCT_BG)
    _net_row("TOTAL DEDUCTIONS (USD)", total_deductions_usd,'"$"#,##0.00', C_DEDUCT_BG, C_DEDUCT_BG)
    _net_row("NET PAYABLE (USD)",      net_usd,             '"$"#,##0.00', "2E7D32", "1B5E20", size=14, height=32)
    _net_row("NET PAYABLE (TL)",       net_tl,              '"₺"#,##0.00', "B71C1C", "B71C1C", size=13, height=28)


# ── Standby detail sheet ───────────────────────────────────────────────────────

def _build_standby_sheet(ws, cname, month, year, standby_entries, rate_s,
                         rig_summary, total_payable_sb_hours, total_standby):
    ws.sheet_view.showGridLines = False
    # 10 columns: # | Rig | Date | Hole | Start | End | Type | Detail | Hours | Amount
    sb_widths = [5, 16, 11, 14, 9, 9, 18, 28, 9, 14]
    for i, w in enumerate(sb_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    SBCOLS = 10

    row = 1
    ws.row_dimensions[row].height = 35
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=SBCOLS)
    c = ws.cell(row=row, column=1,
                value=f"STANDBY HOURS — {cname} — {MONTHS[month-1]} {year}")
    c.font = Font(bold=True, size=14, color=C_WHITE)
    c.fill = PatternFill("solid", fgColor=C_STANDBY_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    row += 2

    # ── Full detail table ──────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=SBCOLS)
    c2 = ws.cell(row=row, column=1, value="  STANDBY HOURS DETAIL")
    c2.font = Font(bold=True, color=C_WHITE, size=12)
    c2.fill = PatternFill("solid", fgColor=C_STANDBY_BG)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    c2.border = thin_border()
    row += 1

    for col, h in enumerate([
        "#", "Drill Rig", "Date", "Hole ID", "Start", "End",
        "Type (Bekleme Türü)", "Detail (Detay)", "Hours", "Amount (USD)"
    ], 1):
        _cell(ws, row, col, h, bold=True, fg=C_WHITE, bg=C_STANDBY_BG, align="center")
    row += 1

    for idx, e in enumerate(standby_entries, 1):
        hours  = e["hours"]
        amount = hours * rate_s
        _cell(ws, row, 1, idx, align="center")
        _cell(ws, row, 2, e["rig_name"] if "rig_name" in e.keys() else "", bold=True)
        _cell(ws, row, 3, e["entry_date"],  align="center")
        _cell(ws, row, 4, e["hole_id"])
        _cell(ws, row, 5, e["start_time"], align="center")
        _cell(ws, row, 6, e["end_time"],   align="center")
        _cell(ws, row, 7, e["standby_type"])
        _cell(ws, row, 8, e["description"] or "")
        _cell(ws, row, 9, hours,  align="right", fmt="#,##0.00")
        _cell(ws, row, 10, amount, align="right", fmt='"$"#,##0.00', bold=True)
        row += 1

    if not standby_entries:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=SBCOLS)
        _cell(ws, row, 1, "No standby entries for this period.", italic=True, align="center")
        row += 1

    row += 1  # spacer

    # ── Blasting deduction summary per rig ────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=SBCOLS)
    c3 = ws.cell(row=row, column=1,
                 value="  PATLATMA (BLASTING) DEDUCTION SUMMARY  "
                       "— each rig: 24 free hrs/month")
    c3.font = Font(bold=True, color=C_WHITE, size=12)
    c3.fill = PatternFill("solid", fgColor="B71C1C")
    c3.alignment = Alignment(horizontal="left", vertical="center")
    c3.border = thin_border()
    row += 1

    for col, h in enumerate([
        "Drill Rig", "Non-Blasting (h)", "Patlatma Total (h)",
        "−24h Free", "Patlatma Paid (h)", "Net Payable (h)",
        "Rate / hr (USD)", "Net Amount (USD)", "", ""
    ], 1):
        _cell(ws, row, col, h, bold=True, fg=C_WHITE, bg="B71C1C", align="center")
    row += 1

    for r in rig_summary:
        _cell(ws, row, 1, r["rig"], bold=True, bg="FFEBEE")
        _cell(ws, row, 2, r["other"],          align="right", fmt="#,##0.00", bg="FFEBEE")
        _cell(ws, row, 3, r["blasting_total"],  align="right", fmt="#,##0.00", bg="FFEBEE")
        _cell(ws, row, 4, -r["blasting_free"],  align="right", fmt="#,##0.00", bg="FFEBEE",
              fg="B71C1C")
        _cell(ws, row, 5, r["blasting_paid"],   align="right", fmt="#,##0.00", bg="FFEBEE")
        _cell(ws, row, 6, r["payable_hours"],   align="right", fmt="#,##0.00", bg="FFEBEE",
              bold=True)
        _cell(ws, row, 7, rate_s,               align="right", fmt='"$"#,##0.00', bg="FFEBEE")
        _cell(ws, row, 8, r["amount"],          align="right", fmt='"$"#,##0.00', bg="FFEBEE",
              bold=True)
        _cell(ws, row, 9, "", bg="FFEBEE")
        _cell(ws, row, 10, "", bg="FFEBEE")
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    _cell(ws, row, 1, "TOTAL NET PAYABLE", bold=True, bg="FFE0B2", align="right",
          fg=C_STANDBY_BG)
    _cell(ws, row, 6, total_payable_sb_hours, align="right", fmt="#,##0.00",
          bold=True, bg="FFE0B2")
    _cell(ws, row, 7, "",             bg="FFE0B2")
    _cell(ws, row, 8, total_standby,  align="right", fmt='"$"#,##0.00', bold=True,
          bg="FFE0B2", fg=C_STANDBY_BG)
    _cell(ws, row, 9, "", bg="FFE0B2")
    _cell(ws, row, 10, "", bg="FFE0B2")


# ── PPE / Diesel detail sheet ──────────────────────────────────────────────────

def _build_deductions_sheet(ws, cname, ctype, month, year, charge_rows,
                            total_deductions_tl, total_deductions_usd, usd_tl_rate):
    ws.sheet_view.showGridLines = False
    is_ppe   = ctype == "underground"
    bg_color = C_DEDUCT_BG
    label    = "PPE CHARGES" if is_ppe else "DIESEL CHARGES"

    if is_ppe:
        # 9 columns for PPE: # | Mat.Code | Description | Qty | Unit | Date | Unit Price | Total TL | Total USD
        PPE_NCOLS = 9
        ppe_widths = [5, 16, 30, 10, 10, 13, 16, 16, 16]
        for i, w in enumerate(ppe_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        for i, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    ncols = 9  # both layouts use 9 cols

    row = 1
    ws.row_dimensions[row].height = 35
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1,
                value=f"{label} — {cname} — {MONTHS[month-1]} {year}")
    c.font = Font(bold=True, size=14, color=C_WHITE)
    c.fill = PatternFill("solid", fgColor=bg_color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    row += 2

    if is_ppe:
        row = _section_header(ws, row, f"  {label} DETAIL (in TL)", bg=bg_color)
        # Override NCOLS for section header — already merged to 9
        row = _col_headers(ws, row, [
            "#", "Material Code", "Description (Malzeme kısa metni)",
            "Qty", "Unit", "Date", "Unit Price (TL)", "Total (TL)", "Total (USD)"
        ], bg_color)

        for idx, r in enumerate(charge_rows, 1):
            qty   = r["quantity"]
            price = r["unit_price"]
            total_tl  = qty * price
            total_usd = total_tl / usd_tl_rate if usd_tl_rate else 0.0
            _cell(ws, row, 1, idx, align="center")
            _cell(ws, row, 2, r["material_code"]   if "material_code"   in r.keys() else "")
            _cell(ws, row, 3, r["item_name"])
            _cell(ws, row, 4, qty,   align="right", fmt="#,##0.00")
            _cell(ws, row, 5, r["unit_of_measure"] if "unit_of_measure" in r.keys() else "",
                  align="center")
            _cell(ws, row, 6, r["entry_date"]      if "entry_date"      in r.keys() else "",
                  align="center")
            _cell(ws, row, 7, price,     align="right", fmt='"₺"#,##0.00')
            _cell(ws, row, 8, total_tl,  align="right", fmt='"₺"#,##0.00', bold=True)
            _cell(ws, row, 9, total_usd, align="right", fmt='"$"#,##0.00', bold=True)
            row += 1

        if not charge_rows:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
            _cell(ws, row, 1, "No charges for this period.", italic=True, align="center")
            row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        _cell(ws, row, 1, "TOTAL (TL)", bold=True, bg="FFEBEE", align="right")
        _cell(ws, row, 8, total_deductions_tl,
              align="right", fmt='"₺"#,##0.00', bold=True, bg="FFEBEE", fg=bg_color)
        _cell(ws, row, 9, total_deductions_usd,
              align="right", fmt='"$"#,##0.00', bold=True, bg="FFEBEE", fg=bg_color)
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        _cell(ws, row, 1,
              f"TOTAL (USD)  [÷ {usd_tl_rate:,.4f} TL/USD]",
              bold=True, bg="FFCDD2", align="right")
        _cell(ws, row, 8, "",                bg="FFCDD2")
        _cell(ws, row, 9, total_deductions_usd,
              align="right", fmt='"$"#,##0.00', bold=True, bg="FFCDD2", fg=bg_color)

    else:
        # Diesel — original layout
        row = _section_header(ws, row, f"  {label} DETAIL (in TL)", bg=bg_color)
        row = _col_headers(ws, row, [
            "#", "Item / Description", "", "", "", "",
            "Quantity", "Unit Price (TL)", "Total (TL)"
        ], bg_color)

        for idx, r in enumerate(charge_rows, 1):
            qty   = r["quantity"]
            price = r["unit_price"]
            total = qty * price
            _cell(ws, row, 1, idx, align="center")
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            _cell(ws, row, 2, r["description"])
            _cell(ws, row, 7, qty,   align="right", fmt="#,##0.00")
            _cell(ws, row, 8, price, align="right", fmt='"₺"#,##0.00')
            _cell(ws, row, 9, total, align="right", fmt='"₺"#,##0.00', bold=True)
            row += 1

        if not charge_rows:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
            _cell(ws, row, 1, "No charges for this period.", italic=True, align="center")
            row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        _cell(ws, row, 1, "TOTAL (TL)", bold=True, bg="FFEBEE", align="right")
        _cell(ws, row, 9, total_deductions_tl,
              align="right", fmt='"₺"#,##0.00', bold=True, bg="FFEBEE", fg=bg_color)
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        _cell(ws, row, 1,
              f"TOTAL (USD)  [÷ {usd_tl_rate:,.4f} TL/USD]",
              bold=True, bg="FFCDD2", align="right")
        _cell(ws, row, 9, total_deductions_usd,
              align="right", fmt='"$"#,##0.00', bold=True, bg="FFCDD2", fg=bg_color)


# ── Summary sheet (multi-contractor) ──────────────────────────────────────────

def _build_summary_sheet(wb, summary_data, project, month, year, usd_tl_rate):
    ws = wb.create_sheet(title="Summary", index=0)
    ws.sheet_view.showGridLines = False

    col_widths = [26, 12, 14, 14, 16, 16, 16, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    SCOLS = 8
    row = 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=SCOLS)
    c = ws.cell(row=1, column=1,
                value=f"MONTHLY SUMMARY — {MONTHS[month-1].upper()} {year}")
    c.font = Font(bold=True, size=16, color=C_WHITE)
    c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    row = 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=SCOLS)
    _cell(ws, row, 1,
          f"Project: {project['name']}   |   1 USD = {usd_tl_rate:,.4f} TL",
          bold=True, align="center", bg="ECEFF1")
    row += 2

    hdrs = ["Contractor", "Type", "Boreholes (USD)", "Standby (USD)",
            "Deductions (TL)", "Deductions (USD)", "Net (USD)", "Net (TL)"]
    for col, h in enumerate(hdrs, 1):
        _cell(ws, row, col, h, bold=True, fg=C_WHITE, bg=C_COL_HDR_BG, align="center")
    row += 1

    g_bh = g_sb = g_dtl = g_dusd = g_nusd = g_ntl = 0.0
    for s in summary_data:
        _cell(ws, row, 1, s["name"])
        _cell(ws, row, 2, s["type"].capitalize(), align="center")
        _cell(ws, row, 3, s["boreholes"],   align="right", fmt='"$"#,##0.00')
        _cell(ws, row, 4, s["standby"],     align="right", fmt='"$"#,##0.00')
        _cell(ws, row, 5, s["deduct_tl"],   align="right", fmt='"₺"#,##0.00')
        _cell(ws, row, 6, s["deduct_usd"],  align="right", fmt='"$"#,##0.00')
        _cell(ws, row, 7, s["net_usd"],     align="right", fmt='"$"#,##0.00', bold=True)
        _cell(ws, row, 8, s["net_tl"],      align="right", fmt='"₺"#,##0.00', bold=True,
              fg="B71C1C")
        g_bh   += s["boreholes"]
        g_sb   += s["standby"]
        g_dtl  += s["deduct_tl"]
        g_dusd += s["deduct_usd"]
        g_nusd += s["net_usd"]
        g_ntl  += s["net_tl"]
        row += 1

    _cell(ws, row, 1, "TOTAL", bold=True, bg=C_NET_BG, align="right")
    _cell(ws, row, 2, "", bg=C_NET_BG)
    _cell(ws, row, 3, g_bh,   align="right", fmt='"$"#,##0.00', bold=True, bg=C_NET_BG)
    _cell(ws, row, 4, g_sb,   align="right", fmt='"$"#,##0.00', bold=True, bg=C_NET_BG)
    _cell(ws, row, 5, g_dtl,  align="right", fmt='"₺"#,##0.00', bold=True, bg=C_NET_BG)
    _cell(ws, row, 6, g_dusd, align="right", fmt='"$"#,##0.00', bold=True, bg=C_NET_BG)
    _cell(ws, row, 7, g_nusd, align="right", fmt='"$"#,##0.00', bold=True, bg=C_NET_BG,
          fg="1B5E20")
    _cell(ws, row, 8, g_ntl,  align="right", fmt='"₺"#,##0.00', bold=True, bg=C_NET_BG,
          fg="B71C1C")
