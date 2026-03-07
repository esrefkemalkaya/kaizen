from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QLabel, QComboBox, QMessageBox, QSplitter,
    QFileDialog, QLineEdit, QStyledItemDelegate, QCompleter, QApplication
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QKeySequence, QShortcut
from datetime import datetime, timedelta

import db.models as m
from .styles import (PAGE_TITLE_STYLE, TABLE_STYLE, BTN_PRIMARY,
                     BTN_DANGER, BTN_SUCCESS, COMBO_STYLE, LABEL_MUTED)

MONTHS = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]

STANDBY_TYPES = [
    "Patlatma", "Elektrik arızası", "Su kesintisi",
    "Topograf beklendi", "Servis ekibi beklendi", "Hava koşulları",
    "Bakım", "Diğer",
]

# Standby table column indices (matches Excel order)
SB_COL_RIG   = 0   # Drill Rig
SB_COL_HOLE  = 1   # Hole ID
SB_COL_DATE  = 2   # Tarih
SB_COL_START = 3   # Start
SB_COL_END   = 4   # Finish
SB_COL_TYPE  = 5   # Bekleme Türü
SB_COL_DESC  = 6   # Detay
SB_COL_HOURS = 7   # Bekleme(saat) — read-only, computed
SB_COL_AMT   = 8   # Amount — read-only, computed

SB_HEADERS = [
    "Drill Rig", "Hole ID", "Date (Tarih)", "Start", "Finish (End)",
    "Type (Bekleme Türü)", "Detail (Detay)", "Hours", "Amount"
]

# Excel header → app column index map (partial, case-insensitive)
SB_EXCEL_MAP = {
    "drill rig":     SB_COL_RIG,
    "rig":           SB_COL_RIG,
    "hole id":       SB_COL_HOLE,
    "hole":          SB_COL_HOLE,
    "tarih":         SB_COL_DATE,
    "date":          SB_COL_DATE,
    "start":         SB_COL_START,
    "finish":        SB_COL_END,
    "end":           SB_COL_END,
    "bekleme türü":  SB_COL_TYPE,
    "bekleme turu":  SB_COL_TYPE,
    "tür":           SB_COL_TYPE,
    "tur":           SB_COL_TYPE,
    "detay":         SB_COL_DESC,
    "detail":        SB_COL_DESC,
    "bekleme(saat)": SB_COL_HOURS,
    "bekleme":       SB_COL_HOURS,
    "saat":          SB_COL_HOURS,
}


def _match_sb_col(header: str) -> int:
    h = header.strip().lower()
    for key, idx in SB_EXCEL_MAP.items():
        if key in h or h in key:
            return idx
    return -1  # unknown


MACHINES = ["GEO 900E-1", "GEO 900E-2", "GEO 900E-3", "GEO 900E-5"]

# Borehole table column indices
BH_COL_RIG         = 0   # Makine No
BH_COL_HOLE        = 1   # Hole ID
BH_COL_START_DATE  = 2   # Start Date
BH_COL_END_DATE    = 3   # End Date
BH_COL_START_DEPTH = 4   # Metre Başlangıç
BH_COL_END_DEPTH   = 5   # Metre Bitiş
BH_COL_METERS      = 6   # İlerleme (m)
BH_COL_AMOUNT      = 7   # Amount

BH_HEADERS = [
    "Makine No", "Hole ID", "Start Date", "End Date",
    "M. Başlangıç", "M. Bitiş", "İlerleme (m)", "Amount"
]


def _calc_hours(start: str, end: str) -> float:
    for fmt in ["%H:%M", "%H.%M", "%H:%M:%S"]:
        try:
            s = datetime.strptime(start.strip(), fmt)
            e = datetime.strptime(end.strip(), fmt)
            diff = e - s
            if diff.total_seconds() < 0:
                diff += timedelta(days=1)
            return round(diff.total_seconds() / 3600, 2)
        except ValueError:
            continue
    return 0.0


class ComboDelegate(QStyledItemDelegate):
    """Dropdown delegate for a fixed list of choices."""
    def __init__(self, choices: list[str], parent=None):
        super().__init__(parent)
        self._choices = choices

    def createEditor(self, parent, option, index):
        combo = QComboBox(parent)
        combo.addItems(self._choices)
        return combo

    def setEditorData(self, editor, index):
        val = index.data() or ""
        idx = editor.findText(val)
        if idx >= 0:
            editor.setCurrentIndex(idx)

    def setModelData(self, editor, model, index):
        model.setData(index, editor.currentText(), Qt.ItemDataRole.EditRole)

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect)


class AutoCompleteDelegate(QStyledItemDelegate):
    def __init__(self, items: list[str] | None = None, parent=None):
        super().__init__(parent)
        self._items: list[str] = items or []

    def set_completions(self, items: list[str]):
        self._items = items

    def createEditor(self, parent, option, index):
        editor = QLineEdit(parent)
        completer = QCompleter(self._items, editor)
        completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        completer.setFilterMode(Qt.MatchFlag.MatchContains)
        editor.setCompleter(completer)
        return editor

    def setEditorData(self, editor, index):
        editor.setText(index.data() or "")

    def setModelData(self, editor, model, index):
        model.setData(index, editor.text(), Qt.ItemDataRole.EditRole)


class DrillingEntryView(QWidget):
    def __init__(self):
        super().__init__()
        self._project_id = None
        self._project_name = ""
        self._contractor_id = None
        self._borehole_ids: list[int | None] = []
        self._rig_delegate = ComboDelegate(MACHINES)
        self._standby_ids: list[int | None] = []
        self._hole_delegate = AutoCompleteDelegate()
        self._sb_hole_delegate = AutoCompleteDelegate()
        self._type_delegate = AutoCompleteDelegate(STANDBY_TYPES)
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(14)

        title = QLabel("Drilling Entries")
        title.setStyleSheet(PAGE_TITLE_STYLE)
        layout.addWidget(title)

        self.ctx_label = QLabel("No project selected.")
        self.ctx_label.setStyleSheet(LABEL_MUTED)
        layout.addWidget(self.ctx_label)

        # ── Filter bar ───────────────────────────────────────────────────
        filter_bar = QHBoxLayout()
        filter_bar.setSpacing(12)

        filter_bar.addWidget(QLabel("Contractor:"))
        self.contractor_combo = QComboBox()
        self.contractor_combo.setStyleSheet(COMBO_STYLE)
        self.contractor_combo.setMinimumWidth(200)
        self.contractor_combo.currentIndexChanged.connect(self._on_contractor_changed)
        filter_bar.addWidget(self.contractor_combo)

        filter_bar.addWidget(QLabel("Month:"))
        self.month_combo = QComboBox()
        self.month_combo.setStyleSheet(COMBO_STYLE)
        for i, name in enumerate(MONTHS):
            self.month_combo.addItem(name, i + 1)
        filter_bar.addWidget(self.month_combo)

        filter_bar.addWidget(QLabel("Year:"))
        self.year_combo = QComboBox()
        self.year_combo.setStyleSheet(COMBO_STYLE)
        from datetime import date
        current_year = date.today().year
        for y in range(current_year - 2, current_year + 3):
            self.year_combo.addItem(str(y), y)
        self.year_combo.setCurrentText(str(current_year))
        filter_bar.addWidget(self.year_combo)

        load_btn = QPushButton("Load")
        load_btn.setStyleSheet(BTN_PRIMARY)
        load_btn.clicked.connect(self._load_all)
        filter_bar.addWidget(load_btn)
        filter_bar.addStretch()
        layout.addLayout(filter_bar)

        # ── Splitter (Boreholes top / Standby bottom) ────────────────────
        splitter = QSplitter(Qt.Orientation.Vertical)

        # ── Boreholes section ─────────────────────────────────────────────
        bh_widget = QWidget()
        bh_layout = QVBoxLayout(bh_widget)
        bh_layout.setContentsMargins(0, 4, 0, 4)
        bh_layout.setSpacing(6)

        bh_label = QLabel("Boreholes")
        bh_label.setStyleSheet("font-weight: bold; font-size: 13px; color: #1565c0;")
        bh_layout.addWidget(bh_label)

        bh_action = QHBoxLayout()
        self.bh_add_btn = QPushButton("+ Add Row")
        self.bh_add_btn.setStyleSheet(BTN_PRIMARY)
        self.bh_add_btn.clicked.connect(self._bh_add_row)
        self.bh_del_btn = QPushButton("Delete Row")
        self.bh_del_btn.setStyleSheet(BTN_DANGER)
        self.bh_del_btn.clicked.connect(self._bh_delete_row)
        self.bh_save_btn = QPushButton("Save Boreholes")
        self.bh_save_btn.setStyleSheet(BTN_SUCCESS)
        self.bh_save_btn.clicked.connect(self._bh_save)
        self.bh_import_btn = QPushButton("Import from Excel")
        self.bh_import_btn.setStyleSheet(BTN_PRIMARY)
        self.bh_import_btn.clicked.connect(self._import_excel)
        for btn in [self.bh_add_btn, self.bh_del_btn, self.bh_save_btn, self.bh_import_btn]:
            bh_action.addWidget(btn)
        bh_action.addStretch()
        bh_layout.addLayout(bh_action)

        # 8 cols: Makine No | Hole ID | Start Date | End Date | M.Başlangıç | M.Bitiş | İlerleme | Amount
        self.bh_table = QTableWidget()
        self.bh_table.setStyleSheet(TABLE_STYLE)
        self.bh_table.setColumnCount(8)
        self.bh_table.setHorizontalHeaderLabels(BH_HEADERS)
        self.bh_table.setColumnWidth(BH_COL_RIG,         120)
        self.bh_table.horizontalHeader().setSectionResizeMode(BH_COL_HOLE, QHeaderView.ResizeMode.Stretch)
        self.bh_table.setColumnWidth(BH_COL_START_DATE,   95)
        self.bh_table.setColumnWidth(BH_COL_END_DATE,     95)
        self.bh_table.setColumnWidth(BH_COL_START_DEPTH, 100)
        self.bh_table.setColumnWidth(BH_COL_END_DEPTH,   100)
        self.bh_table.setColumnWidth(BH_COL_METERS,       85)
        self.bh_table.setColumnWidth(BH_COL_AMOUNT,      110)
        self.bh_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.bh_table.verticalHeader().setVisible(False)
        self.bh_table.setItemDelegateForColumn(BH_COL_RIG,  self._rig_delegate)
        self.bh_table.setItemDelegateForColumn(BH_COL_HOLE, self._hole_delegate)
        self.bh_table.itemChanged.connect(self._bh_on_item_changed)
        bh_layout.addWidget(self.bh_table)

        bh_footer = QHBoxLayout()
        self.bh_meters_lbl = QLabel("Total Meters: 0.00")
        self.bh_amount_lbl = QLabel("Total: $0.00")
        self.bh_amount_lbl.setStyleSheet("font-weight: bold; color: #1b5e20;")
        bh_footer.addWidget(self.bh_meters_lbl)
        bh_footer.addStretch()
        bh_footer.addWidget(self.bh_amount_lbl)
        bh_layout.addLayout(bh_footer)

        splitter.addWidget(bh_widget)

        # ── Standby section ───────────────────────────────────────────────
        sb_widget = QWidget()
        sb_layout = QVBoxLayout(sb_widget)
        sb_layout.setContentsMargins(0, 4, 0, 4)
        sb_layout.setSpacing(6)

        sb_label = QLabel("Standby Hours")
        sb_label.setStyleSheet("font-weight: bold; font-size: 13px; color: #e65100;")
        sb_layout.addWidget(sb_label)

        sb_action = QHBoxLayout()
        self.sb_add_btn = QPushButton("+ Add Row")
        self.sb_add_btn.setStyleSheet(BTN_PRIMARY)
        self.sb_add_btn.clicked.connect(self._sb_add_row)
        self.sb_del_btn = QPushButton("Delete Row")
        self.sb_del_btn.setStyleSheet(BTN_DANGER)
        self.sb_del_btn.clicked.connect(self._sb_delete_row)
        self.sb_import_btn = QPushButton("Import Excel…")
        self.sb_import_btn.setStyleSheet(BTN_PRIMARY)
        self.sb_import_btn.clicked.connect(self._sb_import_excel)
        self.sb_paste_btn = QPushButton("Paste from Excel (Ctrl+V)")
        self.sb_paste_btn.setStyleSheet(BTN_PRIMARY)
        self.sb_paste_btn.clicked.connect(self._sb_paste_clipboard)
        self.sb_save_btn = QPushButton("Save Standby")
        self.sb_save_btn.setStyleSheet(BTN_SUCCESS)
        self.sb_save_btn.clicked.connect(self._sb_save)
        for btn in [self.sb_add_btn, self.sb_del_btn,
                    self.sb_import_btn, self.sb_paste_btn, self.sb_save_btn]:
            sb_action.addWidget(btn)
        sb_action.addStretch()
        sb_layout.addLayout(sb_action)

        hint = QLabel("Tip: Copy rows from your standby Excel sheet then click 'Paste from Excel' or Ctrl+V.")
        hint.setStyleSheet("color: #607d8b; font-size: 11px;")
        sb_layout.addWidget(hint)

        # 9 cols matching Excel: Rig | Hole | Date | Start | End | Type | Detail | Hours | Amount
        self.sb_table = QTableWidget()
        self.sb_table.setStyleSheet(TABLE_STYLE)
        self.sb_table.setColumnCount(9)
        self.sb_table.setHorizontalHeaderLabels(SB_HEADERS)
        self.sb_table.setColumnWidth(SB_COL_RIG,   110)
        self.sb_table.setColumnWidth(SB_COL_HOLE,  110)
        self.sb_table.setColumnWidth(SB_COL_DATE,   90)
        self.sb_table.setColumnWidth(SB_COL_START,  65)
        self.sb_table.setColumnWidth(SB_COL_END,    65)
        self.sb_table.setColumnWidth(SB_COL_TYPE,  130)
        self.sb_table.horizontalHeader().setSectionResizeMode(SB_COL_DESC, QHeaderView.ResizeMode.Stretch)
        self.sb_table.setColumnWidth(SB_COL_HOURS,  65)
        self.sb_table.setColumnWidth(SB_COL_AMT,   110)
        self.sb_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.sb_table.verticalHeader().setVisible(False)
        self.sb_table.setItemDelegateForColumn(SB_COL_HOLE, self._sb_hole_delegate)
        self.sb_table.setItemDelegateForColumn(SB_COL_TYPE, self._type_delegate)
        self.sb_table.itemChanged.connect(self._sb_on_item_changed)
        sb_layout.addWidget(self.sb_table)

        # Ctrl+V shortcut on standby table
        sb_sc = QShortcut(QKeySequence("Ctrl+V"), self.sb_table)
        sb_sc.activated.connect(self._sb_paste_clipboard)

        sb_footer = QHBoxLayout()
        self.sb_hours_lbl = QLabel("Total Hours: 0.00")
        self.sb_blast_lbl = QLabel("")
        self.sb_blast_lbl.setStyleSheet("color: #b71c1c; font-size: 11px;")
        self.sb_amount_lbl = QLabel("Net Payable: $0.00")
        self.sb_amount_lbl.setStyleSheet("font-weight: bold; color: #e65100;")
        sb_footer.addWidget(self.sb_hours_lbl)
        sb_footer.addWidget(self.sb_blast_lbl)
        sb_footer.addStretch()
        sb_footer.addWidget(self.sb_amount_lbl)
        sb_layout.addLayout(sb_footer)

        splitter.addWidget(sb_widget)
        splitter.setSizes([380, 340])
        layout.addWidget(splitter)

    # ── Public API ────────────────────────────────────────────────────────────

    def set_project(self, project_id: int, project_name: str):
        self._project_id = project_id
        self._project_name = project_name
        self.ctx_label.setText(f"Project: {project_name}")
        self._refresh_contractors()

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _refresh_contractors(self):
        self.contractor_combo.blockSignals(True)
        self.contractor_combo.clear()
        if self._project_id:
            for c in m.get_contractors(self._project_id):
                self.contractor_combo.addItem(
                    f"{c['name']} ({c['type'].capitalize()})", c["id"]
                )
        self.contractor_combo.blockSignals(False)
        self._load_all()

    def _on_contractor_changed(self):
        self._contractor_id = self.contractor_combo.currentData()

    def _get_rate(self):
        cid = self.contractor_combo.currentData()
        if cid:
            c = m.get_contractor(cid)
            if c:
                return c["rate_per_meter"], c["standby_hour_rate"]
        return 0.0, 0.0

    def _refresh_completions(self):
        cid = self.contractor_combo.currentData()
        if cid:
            hole_ids = m.get_all_hole_ids(cid)
            self._hole_delegate.set_completions(hole_ids)
            all_holes = m.get_all_hole_ids_for_standby(cid)
            self._sb_hole_delegate.set_completions(all_holes)

    def _load_all(self):
        self._load_boreholes()
        self._load_standby()

    # ── Borehole table ────────────────────────────────────────────────────────

    def _load_boreholes(self):
        cid = self.contractor_combo.currentData()
        month = self.month_combo.currentData()
        year = self.year_combo.currentData()
        if not cid or not month or not year:
            return
        self._contractor_id = cid
        self._refresh_completions()
        entries = m.get_drilling_entries(cid, month, year)
        rate_m, _ = self._get_rate()
        self._borehole_ids = []
        self.bh_table.blockSignals(True)
        self.bh_table.setRowCount(0)
        for e in entries:
            self._bh_append_row(
                e["id"], e["rig_name"], e["hole_id"],
                e["start_date"], e["end_date"],
                e["start_depth"], e["end_depth"],
                e["meters_drilled"], rate_m
            )
        self.bh_table.blockSignals(False)
        self._bh_update_totals()

    def _bh_append_row(self, db_id, rig_name, hole_id, start_date, end_date,
                       start_depth, end_depth, meters, rate_m):
        row = self.bh_table.rowCount()
        self.bh_table.insertRow(row)
        self._borehole_ids.append(db_id)

        self.bh_table.setItem(row, BH_COL_RIG,         QTableWidgetItem(str(rig_name or "")))
        self.bh_table.setItem(row, BH_COL_HOLE,        QTableWidgetItem(str(hole_id or "")))
        self.bh_table.setItem(row, BH_COL_START_DATE,  QTableWidgetItem(str(start_date or "")))
        self.bh_table.setItem(row, BH_COL_END_DATE,    QTableWidgetItem(str(end_date or "")))
        self.bh_table.setItem(row, BH_COL_START_DEPTH, QTableWidgetItem(f"{start_depth:.2f}"))
        self.bh_table.setItem(row, BH_COL_END_DEPTH,   QTableWidgetItem(f"{end_depth:.2f}"))
        self.bh_table.setItem(row, BH_COL_METERS,      QTableWidgetItem(f"{meters:.2f}"))

        amt_item = QTableWidgetItem(f"${meters * rate_m:,.2f}")
        amt_item.setFlags(amt_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        amt_item.setForeground(QColor("#1b5e20"))
        self.bh_table.setItem(row, BH_COL_AMOUNT, amt_item)

    def _bh_add_row(self):
        if not self._contractor_id:
            QMessageBox.information(self, "Select", "Select a contractor first.")
            return
        rate_m, _ = self._get_rate()
        self.bh_table.blockSignals(True)
        self._bh_append_row(None, MACHINES[0], "", "", "", 0, 0, 0, rate_m)
        self.bh_table.blockSignals(False)

    def _bh_delete_row(self):
        row = self.bh_table.currentRow()
        if row < 0:
            return
        db_id = self._borehole_ids[row]
        if db_id:
            reply = QMessageBox.question(
                self, "Confirm", "Delete this row?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
            m.delete_drilling_entry(db_id)
        self.bh_table.removeRow(row)
        self._borehole_ids.pop(row)
        self._bh_update_totals()

    def _bh_cell(self, row, col) -> str:
        it = self.bh_table.item(row, col)
        return it.text().strip() if it else ""

    def _bh_save(self):
        cid = self.contractor_combo.currentData()
        month = self.month_combo.currentData()
        year = self.year_combo.currentData()
        if not cid:
            QMessageBox.information(self, "Select", "Select a contractor first.")
            return
        errors = []
        for row in range(self.bh_table.rowCount()):
            rig_name   = self._bh_cell(row, BH_COL_RIG)
            hole_id    = self._bh_cell(row, BH_COL_HOLE)
            if not hole_id:
                errors.append(f"Row {row + 1}: Hole ID is empty.")
                continue
            start_date = self._bh_cell(row, BH_COL_START_DATE)
            end_date   = self._bh_cell(row, BH_COL_END_DATE)
            try:
                start_depth = float(self._bh_cell(row, BH_COL_START_DEPTH) or "0")
                end_depth   = float(self._bh_cell(row, BH_COL_END_DEPTH)   or "0")
                meters      = float(self._bh_cell(row, BH_COL_METERS)      or "0")
            except ValueError:
                errors.append(f"Row {row + 1}: Invalid numeric value.")
                continue
            db_id = self._borehole_ids[row]
            new_id = m.upsert_drilling_entry(
                db_id, cid, month, year,
                hole_id, start_date, end_date, start_depth, end_depth, meters,
                rig_name=rig_name,
            )
            self._borehole_ids[row] = new_id
        if errors:
            QMessageBox.warning(self, "Validation Errors", "\n".join(errors))
        else:
            QMessageBox.information(self, "Saved", "Boreholes saved successfully.")
        self._load_boreholes()

    def _bh_on_item_changed(self, item):
        row = item.row()
        col = item.column()
        if col in (BH_COL_START_DEPTH, BH_COL_END_DEPTH):
            # depth changed → auto-calc meters
            try:
                s = float(self._bh_cell(row, BH_COL_START_DEPTH) or "0")
                e = float(self._bh_cell(row, BH_COL_END_DEPTH)   or "0")
                meters = max(0.0, e - s)
            except ValueError:
                return
            self.bh_table.blockSignals(True)
            self.bh_table.setItem(row, BH_COL_METERS, QTableWidgetItem(f"{meters:.2f}"))
            self._bh_set_amount(row, meters)
            self.bh_table.blockSignals(False)
            self._bh_update_totals()
        elif col == BH_COL_METERS:
            # meters manually edited → update amount
            try:
                meters = float(self._bh_cell(row, BH_COL_METERS) or "0")
            except ValueError:
                return
            self.bh_table.blockSignals(True)
            self._bh_set_amount(row, meters)
            self.bh_table.blockSignals(False)
            self._bh_update_totals()

    def _bh_set_amount(self, row, meters):
        rate_m, _ = self._get_rate()
        amt = QTableWidgetItem(f"${meters * rate_m:,.2f}")
        amt.setFlags(amt.flags() & ~Qt.ItemFlag.ItemIsEditable)
        amt.setForeground(QColor("#1b5e20"))
        self.bh_table.setItem(row, BH_COL_AMOUNT, amt)

    def _bh_update_totals(self):
        rate_m, _ = self._get_rate()
        total_m = total_amt = 0.0
        for row in range(self.bh_table.rowCount()):
            try:
                meters = float(self._bh_cell(row, BH_COL_METERS) or "0")
            except ValueError:
                continue
            total_m += meters
            total_amt += meters * rate_m
        self.bh_meters_lbl.setText(f"Total Meters: {total_m:,.2f}")
        self.bh_amount_lbl.setText(f"Total: ${total_amt:,.2f}")

    def _import_excel(self):
        """Import boreholes from Excel (Giriş sheet format or simple Hole/Meter columns)."""
        if not self._contractor_id:
            QMessageBox.information(self, "Select", "Select a contractor first.")
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "Import Excel File", "", "Excel Files (*.xlsx *.xls *.xlsm)"
        )
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            # Try "Giriş" sheet first, then active
            ws = wb["Giriş"] if "Giriş" in wb.sheetnames else wb.active
            rate_m, _ = self._get_rate()
            imported = 0
            self.bh_table.blockSignals(True)
            # Detect if this is the Giriş format by checking A4 header
            a4 = str(ws["A4"].value or "").strip().lower()
            if "makine" in a4 or "machine" in a4:
                # Giriş format: rows 5..24, A=Rig, B=Hole, C=Start, D=End, E=M.Start, F=M.End
                for r in range(5, 25):
                    rig   = str(ws.cell(r, 1).value or "").strip()
                    hole  = str(ws.cell(r, 2).value or "").strip()
                    if not rig and not hole:
                        continue
                    sd = str(ws.cell(r, 3).value or "").strip()
                    ed = str(ws.cell(r, 4).value or "").strip()
                    try:
                        ms = float(ws.cell(r, 5).value or 0)
                        me = float(ws.cell(r, 6).value or 0)
                    except (ValueError, TypeError):
                        ms = me = 0.0
                    meters = max(0.0, me - ms)
                    self._bh_append_row(None, rig, hole, sd, ed, ms, me, meters, rate_m)
                    imported += 1
            else:
                # Generic: find Hole and Meter columns
                hdr_row = col_hole = col_rig = col_meters = None
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            val = cell.value.strip().lower()
                            if "hole" in val or "kuyu" in val:
                                col_hole = cell.column; hdr_row = cell.row
                            elif "makine" in val or "rig" in val:
                                col_rig = cell.column
                            elif "meter" in val or "metre" in val or "ilerleme" in val:
                                col_meters = cell.column
                    if hdr_row:
                        break
                if not hdr_row or not col_hole:
                    QMessageBox.warning(self, "Import Error",
                        "Could not find required columns (Hole/Kuyu).")
                    self.bh_table.blockSignals(False)
                    return
                for row in ws.iter_rows(min_row=hdr_row + 1):
                    hole_val = row[col_hole - 1].value
                    if hole_val is None:
                        continue
                    rig_val = str(row[col_rig - 1].value or "").strip() if col_rig else ""
                    meters_val = row[col_meters - 1].value if col_meters else None
                    try:
                        meters = float(meters_val) if meters_val is not None else 0.0
                    except (ValueError, TypeError):
                        continue
                    self._bh_append_row(None, rig_val, str(hole_val).strip(), "", "", 0, 0, meters, rate_m)
                    imported += 1
            self.bh_table.blockSignals(False)
            self._bh_update_totals()
            QMessageBox.information(self, "Import Complete",
                f"Imported {imported} rows. Click Save Boreholes to keep them.")
        except Exception as e:
            QMessageBox.critical(self, "Import Failed", str(e))

    # ── Standby table ─────────────────────────────────────────────────────────

    def _load_standby(self):
        cid = self.contractor_combo.currentData()
        month = self.month_combo.currentData()
        year = self.year_combo.currentData()
        if not cid or not month or not year:
            return
        _, rate_s = self._get_rate()
        entries = m.get_standby_entries(cid, month, year)
        self._standby_ids = []
        self.sb_table.blockSignals(True)
        self.sb_table.setRowCount(0)
        for e in entries:
            self._sb_append_row(
                e["id"],
                e["rig_name"],
                e["hole_id"],
                e["entry_date"],
                e["start_time"],
                e["end_time"],
                e["standby_type"],
                e["description"],
                e["hours"],
                rate_s,
            )
        self.sb_table.blockSignals(False)
        self._sb_update_totals()

    def _sb_append_row(self, db_id, rig_name, hole_id, entry_date,
                       start_time, end_time, standby_type, description, hours, rate_s):
        row = self.sb_table.rowCount()
        self.sb_table.insertRow(row)
        self._standby_ids.append(db_id)

        self.sb_table.setItem(row, SB_COL_RIG,   QTableWidgetItem(str(rig_name or "")))
        self.sb_table.setItem(row, SB_COL_HOLE,  QTableWidgetItem(str(hole_id or "")))
        self.sb_table.setItem(row, SB_COL_DATE,  QTableWidgetItem(str(entry_date or "")))
        self.sb_table.setItem(row, SB_COL_START, QTableWidgetItem(str(start_time or "")))
        self.sb_table.setItem(row, SB_COL_END,   QTableWidgetItem(str(end_time or "")))
        self.sb_table.setItem(row, SB_COL_TYPE,  QTableWidgetItem(str(standby_type or "")))
        self.sb_table.setItem(row, SB_COL_DESC,  QTableWidgetItem(str(description or "")))

        hrs_item = QTableWidgetItem(f"{hours:.2f}")
        hrs_item.setFlags(hrs_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.sb_table.setItem(row, SB_COL_HOURS, hrs_item)

        amt_item = QTableWidgetItem(f"${hours * rate_s:,.2f}")
        amt_item.setFlags(amt_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        amt_item.setForeground(QColor("#e65100"))
        self.sb_table.setItem(row, SB_COL_AMT, amt_item)

    def _sb_add_row(self):
        if not self._contractor_id:
            QMessageBox.information(self, "Select", "Select a contractor first.")
            return
        _, rate_s = self._get_rate()
        self.sb_table.blockSignals(True)
        self._sb_append_row(None, "", "", "", "", "", "", "", 0, rate_s)
        self.sb_table.blockSignals(False)

    def _sb_delete_row(self):
        row = self.sb_table.currentRow()
        if row < 0:
            return
        db_id = self._standby_ids[row]
        if db_id:
            reply = QMessageBox.question(
                self, "Confirm", "Delete this row?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
            m.delete_standby_entry(db_id)
        self.sb_table.removeRow(row)
        self._standby_ids.pop(row)
        self._sb_update_totals()

    def _sb_save(self):
        cid = self.contractor_combo.currentData()
        month = self.month_combo.currentData()
        year = self.year_combo.currentData()
        if not cid:
            QMessageBox.information(self, "Select", "Select a contractor first.")
            return
        errors = []
        for row in range(self.sb_table.rowCount()):
            rig_name     = self._sb_cell(row, SB_COL_RIG)
            hole_id      = self._sb_cell(row, SB_COL_HOLE)
            entry_date   = self._sb_cell(row, SB_COL_DATE)
            start_time   = self._sb_cell(row, SB_COL_START)
            end_time     = self._sb_cell(row, SB_COL_END)
            standby_type = self._sb_cell(row, SB_COL_TYPE)
            description  = self._sb_cell(row, SB_COL_DESC)

            if not start_time and not end_time and not description:
                errors.append(f"Row {row + 1}: Start/End time or description required.")
                continue

            # Use pre-computed hours if start/end unavailable (e.g. imported)
            if start_time and end_time:
                hours = _calc_hours(start_time, end_time)
            else:
                try:
                    hours = float(self._sb_cell(row, SB_COL_HOURS) or 0)
                except ValueError:
                    hours = 0.0

            db_id = self._standby_ids[row]
            new_id = m.upsert_standby_entry(
                db_id, cid, month, year,
                entry_date, hole_id, start_time, end_time,
                standby_type, description, hours,
                rig_name=rig_name,
            )
            self._standby_ids[row] = new_id

        if errors:
            QMessageBox.warning(self, "Validation Errors", "\n".join(errors))
        else:
            QMessageBox.information(self, "Saved", "Standby entries saved successfully.")
        self._load_standby()

    def _sb_cell(self, row, col) -> str:
        it = self.sb_table.item(row, col)
        return it.text().strip() if it else ""

    def _sb_on_item_changed(self, item):
        row = item.row()
        col = item.column()
        if col not in (SB_COL_START, SB_COL_END):
            return
        start = self._sb_cell(row, SB_COL_START)
        end   = self._sb_cell(row, SB_COL_END)
        hours = _calc_hours(start, end) if start and end else 0.0
        _, rate_s = self._get_rate()

        self.sb_table.blockSignals(True)
        hrs_item = QTableWidgetItem(f"{hours:.2f}")
        hrs_item.setFlags(hrs_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.sb_table.setItem(row, SB_COL_HOURS, hrs_item)

        amt = QTableWidgetItem(f"${hours * rate_s:,.2f}")
        amt.setFlags(amt.flags() & ~Qt.ItemFlag.ItemIsEditable)
        amt.setForeground(QColor("#e65100"))
        self.sb_table.setItem(row, SB_COL_AMT, amt)
        self.sb_table.blockSignals(False)
        self._sb_update_totals()

    def _sb_update_totals(self):
        _, rate_s = self._get_rate()

        # Build fake entry dicts from table for blasting calc
        entries = []
        for row in range(self.sb_table.rowCount()):
            try:
                hours = float(self._sb_cell(row, SB_COL_HOURS) or 0)
            except ValueError:
                hours = 0.0
            entries.append({
                "rig_name":     self._sb_cell(row, SB_COL_RIG) or "Unknown",
                "standby_type": self._sb_cell(row, SB_COL_TYPE),
                "hours":        hours,
            })

        rig_summary, total_payable_hours, total_amount = m.calc_standby_rig_summary(entries, rate_s)
        total_raw = sum(e["hours"] for e in entries)

        self.sb_hours_lbl.setText(f"Raw Total: {total_raw:,.2f} h")

        # Build blasting deduction info string
        blast_parts = []
        for r in rig_summary:
            if r["blasting_total"] > 0:
                blast_parts.append(
                    f"{r['rig']}: {r['blasting_total']:.2f}h blast "
                    f"(−{r['blasting_free']:.2f} free, {r['blasting_paid']:.2f} paid)"
                )
        if blast_parts:
            self.sb_blast_lbl.setText("  |  Blasting: " + "  |  ".join(blast_parts))
        else:
            self.sb_blast_lbl.setText("")

        self.sb_amount_lbl.setText(
            f"Net Payable: {total_payable_hours:,.2f} h  →  ${total_amount:,.2f}"
        )

    # ── Standby Import from Excel file ────────────────────────────────────────

    def _sb_import_excel(self):
        if not self._contractor_id:
            QMessageBox.information(self, "Select", "Select a contractor first.")
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "Open Standby Excel File", "",
            "Excel Files (*.xlsx *.xls *.xlsm);;All Files (*)"
        )
        if not path:
            return
        try:
            rows = self._sb_read_excel(path)
            self._sb_load_rows(rows)
        except Exception as e:
            QMessageBox.critical(self, "Import Error", str(e))

    def _sb_read_excel(self, path: str) -> list[dict]:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h or "").strip() for h in next(rows_iter)]
        col_map = [_match_sb_col(h) for h in headers]
        result = []
        for raw_row in rows_iter:
            if all(v is None or str(v).strip() == "" for v in raw_row):
                continue
            entry = {}
            for pos, val in enumerate(raw_row):
                if pos >= len(col_map):
                    break
                app_col = col_map[pos]
                if app_col < 0:
                    continue
                entry[app_col] = val
            result.append(entry)
        return result

    def _sb_load_rows(self, rows: list[dict]):
        _, rate_s = self._get_rate()
        self.sb_table.blockSignals(True)
        added = 0
        for entry in rows:
            row_idx = self.sb_table.rowCount()
            self.sb_table.insertRow(row_idx)
            self._standby_ids.append(None)
            for col, val in entry.items():
                if col in (SB_COL_HOURS, SB_COL_AMT):
                    continue  # will compute below
                self.sb_table.setItem(row_idx, col, QTableWidgetItem(str(val) if val is not None else ""))
            # Compute hours from start/end or use imported hours
            start = self._sb_cell(row_idx, SB_COL_START)
            end   = self._sb_cell(row_idx, SB_COL_END)
            if start and end:
                hours = _calc_hours(start, end)
            elif SB_COL_HOURS in entry:
                try:
                    hours = float(str(entry[SB_COL_HOURS]).replace(",", "") or 0)
                except ValueError:
                    hours = 0.0
            else:
                hours = 0.0
            hrs_item = QTableWidgetItem(f"{hours:.2f}")
            hrs_item.setFlags(hrs_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.sb_table.setItem(row_idx, SB_COL_HOURS, hrs_item)
            amt_item = QTableWidgetItem(f"${hours * rate_s:,.2f}")
            amt_item.setFlags(amt_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            amt_item.setForeground(QColor("#e65100"))
            self.sb_table.setItem(row_idx, SB_COL_AMT, amt_item)
            added += 1
        self.sb_table.blockSignals(False)
        self._sb_update_totals()
        QMessageBox.information(self, "Imported",
                                f"{added} row(s) added. Click 'Save Standby' to store them.")

    # ── Standby Paste from clipboard ──────────────────────────────────────────

    def _sb_paste_clipboard(self):
        if not self._contractor_id:
            QMessageBox.information(self, "Select", "Select a contractor first.")
            return
        text = QApplication.clipboard().text()
        if not text.strip():
            QMessageBox.information(self, "Empty Clipboard",
                                    "Nothing in clipboard. Copy rows from Excel first.")
            return
        lines = text.strip().split("\n")
        if not lines:
            return

        first = lines[0].lower()
        has_header = any(k in first for k in (
            "rig", "drill", "hole", "tarih", "date", "start", "finish",
            "bekleme", "detay", "saat"
        ))
        if has_header:
            col_map = [_match_sb_col(h) for h in lines[0].split("\t")]
            data_lines = lines[1:]
        else:
            # Assume Excel column order: Rig | Hole | Date | Start | Finish | Type | Detail | Hours
            col_map = [SB_COL_RIG, SB_COL_HOLE, SB_COL_DATE,
                       SB_COL_START, SB_COL_END, SB_COL_TYPE, SB_COL_DESC, SB_COL_HOURS]
            data_lines = lines

        _, rate_s = self._get_rate()
        self.sb_table.blockSignals(True)
        added = 0
        for line in data_lines:
            cells = line.rstrip("\r").split("\t")
            if all(c.strip() == "" for c in cells):
                continue
            row_idx = self.sb_table.rowCount()
            self.sb_table.insertRow(row_idx)
            self._standby_ids.append(None)
            imported_hours = None
            for pos, cell_val in enumerate(cells):
                if pos >= len(col_map):
                    break
                app_col = col_map[pos]
                if app_col < 0:
                    continue
                if app_col == SB_COL_HOURS:
                    try:
                        imported_hours = float(cell_val.replace(",", "").strip())
                    except ValueError:
                        pass
                    continue  # set below after computing
                if app_col == SB_COL_AMT:
                    continue
                self.sb_table.setItem(row_idx, app_col, QTableWidgetItem(cell_val.strip()))
            # Compute hours
            start = self._sb_cell(row_idx, SB_COL_START)
            end   = self._sb_cell(row_idx, SB_COL_END)
            if start and end:
                hours = _calc_hours(start, end)
            elif imported_hours is not None:
                hours = imported_hours
            else:
                hours = 0.0
            hrs_item = QTableWidgetItem(f"{hours:.2f}")
            hrs_item.setFlags(hrs_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.sb_table.setItem(row_idx, SB_COL_HOURS, hrs_item)
            amt_item = QTableWidgetItem(f"${hours * rate_s:,.2f}")
            amt_item.setFlags(amt_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            amt_item.setForeground(QColor("#e65100"))
            self.sb_table.setItem(row_idx, SB_COL_AMT, amt_item)
            added += 1
        self.sb_table.blockSignals(False)
        self._sb_update_totals()
        QMessageBox.information(self, "Pasted",
                                f"{added} row(s) added. Click 'Save Standby' to store them.")
