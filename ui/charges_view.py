from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QLabel, QComboBox, QMessageBox, QTabWidget, QFileDialog, QApplication
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QKeySequence, QShortcut

import db.models as m
from .styles import (PAGE_TITLE_STYLE, TABLE_STYLE, BTN_PRIMARY,
                     BTN_DANGER, BTN_SUCCESS, COMBO_STYLE, LABEL_MUTED)

MONTHS = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]

# PPE column indices
COL_MAT_CODE = 0
COL_ITEM     = 1
COL_QTY      = 2
COL_UNIT     = 3
COL_DATE     = 4
COL_PRICE    = 5
COL_TOTAL    = 6

PPE_HEADERS = [
    "Material Code", "Description (Malzeme kısa metni)",
    "Qty", "Unit", "Date", "Unit Price (TL)", "Total (TL)"
]

# Excel column name → app column index mapping (case-insensitive, partial match ok)
EXCEL_COL_MAP = {
    "malzeme kısa metni": COL_ITEM,
    "malzeme kisa metni": COL_ITEM,
    "malzeme":            COL_MAT_CODE,
    "miktar":             COL_QTY,
    "giriş ölçü birimi":  COL_UNIT,
    "giris olcu birimi":  COL_UNIT,
    "kayıt tarihi":       COL_DATE,
    "kayit tarihi":       COL_DATE,
    "birim fiyat":        COL_PRICE,
    "tutar":              None,   # computed — skip
}


def _match_excel_col(header: str) -> int | None:
    """Map an Excel column header to app column index. Returns None to skip."""
    h = header.strip().lower()
    for key, idx in EXCEL_COL_MAP.items():
        if key in h or h in key:
            return idx
    return -1  # unknown → try to guess by position


class ChargesView(QWidget):
    def __init__(self):
        super().__init__()
        self._project_id = None
        self._project_name = ""
        self._ppe_row_ids: list[int | None] = []
        self._diesel_row_ids: list[int | None] = []
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(14)

        title = QLabel("Charges (PPE & Diesel)")
        title.setStyleSheet(PAGE_TITLE_STYLE)
        layout.addWidget(title)

        self.ctx_label = QLabel("No project selected.")
        self.ctx_label.setStyleSheet(LABEL_MUTED)
        layout.addWidget(self.ctx_label)

        # ── Filter bar ───────────────────────────────────────────────────
        filter_bar = QHBoxLayout()
        filter_bar.setSpacing(12)

        filter_bar.addWidget(QLabel("Contractor:"))
        self.contractor_combo = QComboBox()
        self.contractor_combo.setStyleSheet(COMBO_STYLE)
        self.contractor_combo.setMinimumWidth(200)
        self.contractor_combo.currentIndexChanged.connect(self._on_contractor_changed)
        filter_bar.addWidget(self.contractor_combo)

        filter_bar.addWidget(QLabel("Month:"))
        self.month_combo = QComboBox()
        self.month_combo.setStyleSheet(COMBO_STYLE)
        for i, name in enumerate(MONTHS):
            self.month_combo.addItem(name, i + 1)
        filter_bar.addWidget(self.month_combo)

        filter_bar.addWidget(QLabel("Year:"))
        self.year_combo = QComboBox()
        self.year_combo.setStyleSheet(COMBO_STYLE)
        from datetime import date
        current_year = date.today().year
        for y in range(current_year - 2, current_year + 3):
            self.year_combo.addItem(str(y), y)
        self.year_combo.setCurrentText(str(current_year))
        filter_bar.addWidget(self.year_combo)

        load_btn = QPushButton("Load")
        load_btn.setStyleSheet(BTN_PRIMARY)
        load_btn.clicked.connect(self._load_charges)
        filter_bar.addWidget(load_btn)
        filter_bar.addStretch()
        layout.addLayout(filter_bar)

        # ── Tabs ─────────────────────────────────────────────────────────
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabBar::tab { padding: 8px 24px; font-size: 13px; }
            QTabBar::tab:selected { background: #1565c0; color: white; font-weight: bold; }
        """)

        self.ppe_widget = self._build_ppe_table()
        self.diesel_widget = self._build_charge_table(
            is_ppe=False,
            col_label="Description",
            placeholder="e.g. Diesel for generator"
        )

        self.tabs.addTab(self.ppe_widget, "PPE Charges (Underground)")
        self.tabs.addTab(self.diesel_widget, "Diesel Charges (Surface)")
        layout.addWidget(self.tabs)

    # ── PPE table (7 columns, import + paste) ─────────────────────────────────

    def _build_ppe_table(self) -> QWidget:
        container = QWidget()
        v = QVBoxLayout(container)
        v.setContentsMargins(0, 12, 0, 0)
        v.setSpacing(10)

        bar = QHBoxLayout()
        add_btn = QPushButton("+ Add Row")
        add_btn.setStyleSheet(BTN_PRIMARY)
        add_btn.clicked.connect(self._add_ppe_row)

        del_btn = QPushButton("Delete Row")
        del_btn.setStyleSheet(BTN_DANGER)
        del_btn.clicked.connect(self._delete_ppe_row)

        import_btn = QPushButton("Import Excel…")
        import_btn.setStyleSheet(BTN_PRIMARY)
        import_btn.clicked.connect(self._import_ppe_excel)

        paste_btn = QPushButton("Paste from Excel (Ctrl+V)")
        paste_btn.setStyleSheet(BTN_PRIMARY)
        paste_btn.clicked.connect(self._paste_ppe_clipboard)

        save_btn = QPushButton("Save All")
        save_btn.setStyleSheet(BTN_SUCCESS)
        save_btn.clicked.connect(self._save_ppe)

        for w in (add_btn, del_btn, import_btn, paste_btn, save_btn):
            bar.addWidget(w)
        bar.addStretch()
        v.addLayout(bar)

        hint = QLabel("Tip: Copy rows from your Excel file then click 'Paste from Excel' or press Ctrl+V while the table is focused.")
        hint.setStyleSheet("color: #607d8b; font-size: 11px;")
        v.addWidget(hint)

        self.ppe_table = QTableWidget()
        self.ppe_table.setStyleSheet(TABLE_STYLE)
        self.ppe_table.setColumnCount(7)
        self.ppe_table.setHorizontalHeaderLabels(PPE_HEADERS)
        hdr = self.ppe_table.horizontalHeader()
        hdr.setSectionResizeMode(COL_ITEM, QHeaderView.ResizeMode.Stretch)
        self.ppe_table.setColumnWidth(COL_MAT_CODE, 120)
        self.ppe_table.setColumnWidth(COL_QTY,      80)
        self.ppe_table.setColumnWidth(COL_UNIT,     80)
        self.ppe_table.setColumnWidth(COL_DATE,     100)
        self.ppe_table.setColumnWidth(COL_PRICE,    120)
        self.ppe_table.setColumnWidth(COL_TOTAL,    120)
        self.ppe_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.ppe_table.verticalHeader().setVisible(False)
        self.ppe_table.itemChanged.connect(
            lambda _: self._recalc_ppe()
        )
        v.addWidget(self.ppe_table)

        # Ctrl+V shortcut on the table
        sc = QShortcut(QKeySequence("Ctrl+V"), self.ppe_table)
        sc.activated.connect(self._paste_ppe_clipboard)

        footer = QHBoxLayout()
        self.ppe_total_lbl = QLabel("Total: ₺0.00")
        self.ppe_total_lbl.setStyleSheet(
            "font-weight: bold; font-size: 14px; color: #c62828; padding: 4px 0;")
        footer.addStretch()
        footer.addWidget(self.ppe_total_lbl)
        v.addLayout(footer)

        return container

    # ── Diesel table (unchanged 4-column layout) ───────────────────────────────

    def _build_charge_table(self, is_ppe: bool, col_label: str, placeholder: str) -> QWidget:
        container = QWidget()
        v = QVBoxLayout(container)
        v.setContentsMargins(0, 12, 0, 0)
        v.setSpacing(10)

        bar = QHBoxLayout()
        add_btn = QPushButton("+ Add Row")
        add_btn.setStyleSheet(BTN_PRIMARY)
        del_btn = QPushButton("Delete Row")
        del_btn.setStyleSheet(BTN_DANGER)
        save_btn = QPushButton("Save All")
        save_btn.setStyleSheet(BTN_SUCCESS)
        bar.addWidget(add_btn)
        bar.addWidget(del_btn)
        bar.addWidget(save_btn)
        bar.addStretch()
        v.addLayout(bar)

        table = QTableWidget()
        table.setStyleSheet(TABLE_STYLE)
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels([col_label, "Quantity", "Unit Price", "Total"])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        table.setColumnWidth(1, 110)
        table.setColumnWidth(2, 130)
        table.setColumnWidth(3, 130)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.verticalHeader().setVisible(False)
        v.addWidget(table)

        footer = QHBoxLayout()
        total_lbl = QLabel("Total: $0.00")
        total_lbl.setStyleSheet(
            "font-weight: bold; font-size: 14px; color: #c62828; padding: 4px 0;")
        footer.addStretch()
        footer.addWidget(total_lbl)
        v.addLayout(footer)

        self.diesel_table = table
        self.diesel_total_lbl = total_lbl
        add_btn.clicked.connect(self._add_diesel_row)
        del_btn.clicked.connect(self._delete_diesel_row)
        save_btn.clicked.connect(self._save_diesel)
        table.itemChanged.connect(lambda _: self._recalc_table(table, total_lbl))

        return container

    # ── Public API ────────────────────────────────────────────────────────────

    def set_project(self, project_id: int, project_name: str):
        self._project_id = project_id
        self._project_name = project_name
        self.ctx_label.setText(f"Project: {project_name}")
        self._refresh_contractors()

    # ── Internal helpers ──────────────────────────────────────────────────────

    def _refresh_contractors(self):
        self.contractor_combo.blockSignals(True)
        self.contractor_combo.clear()
        if self._project_id:
            for c in m.get_contractors(self._project_id):
                self.contractor_combo.addItem(
                    f"{c['name']} ({c['type'].capitalize()})", c["id"]
                )
        self.contractor_combo.blockSignals(False)
        self._load_charges()

    def _on_contractor_changed(self):
        pass  # charges loaded on demand via Load button

    def _load_charges(self):
        cid = self.contractor_combo.currentData()
        month = self.month_combo.currentData()
        year = self.year_combo.currentData()
        if not cid or not month or not year:
            return

        contractor = m.get_contractor(cid)
        if not contractor:
            return
        ctype = contractor["type"]

        # Load PPE
        self._ppe_row_ids = []
        self.ppe_table.blockSignals(True)
        self.ppe_table.setRowCount(0)
        if ctype == "underground":
            for row in m.get_ppe_charges(cid, month, year):
                self._append_ppe_row(
                    row["id"],
                    row["material_code"] if "material_code" in row.keys() else "",
                    row["item_name"],
                    row["quantity"],
                    row["unit_of_measure"] if "unit_of_measure" in row.keys() else "",
                    row["entry_date"] if "entry_date" in row.keys() else "",
                    row["unit_price"],
                )
            self.tabs.setTabEnabled(0, True)
        else:
            self.tabs.setTabEnabled(0, False)
        self.ppe_table.blockSignals(False)
        self._recalc_ppe()

        # Load Diesel
        self._diesel_row_ids = []
        self.diesel_table.blockSignals(True)
        self.diesel_table.setRowCount(0)
        if ctype == "surface":
            for row in m.get_diesel_charges(cid, month, year):
                self._append_charge_row(
                    self.diesel_table, self._diesel_row_ids,
                    row["id"], row["description"], row["quantity"], row["unit_price"]
                )
            self.tabs.setTabEnabled(1, True)
        else:
            self.tabs.setTabEnabled(1, False)
        self.diesel_table.blockSignals(False)
        self._recalc_table(self.diesel_table, self.diesel_total_lbl)

        if ctype == "underground":
            self.tabs.setCurrentIndex(0)
        else:
            self.tabs.setCurrentIndex(1)

    def _append_ppe_row(self, db_id, mat_code, item_name, qty, unit, date_val, unit_price):
        row = self.ppe_table.rowCount()
        self.ppe_table.insertRow(row)
        self._ppe_row_ids.append(db_id)

        self.ppe_table.setItem(row, COL_MAT_CODE, QTableWidgetItem(str(mat_code or "")))
        self.ppe_table.setItem(row, COL_ITEM,     QTableWidgetItem(str(item_name or "")))
        self.ppe_table.setItem(row, COL_QTY,      QTableWidgetItem(str(qty or 0)))
        self.ppe_table.setItem(row, COL_UNIT,     QTableWidgetItem(str(unit or "")))
        self.ppe_table.setItem(row, COL_DATE,     QTableWidgetItem(str(date_val or "")))
        self.ppe_table.setItem(row, COL_PRICE,    QTableWidgetItem(str(unit_price or 0)))

        total = self._parse_float(str(qty)) * self._parse_float(str(unit_price))
        it = QTableWidgetItem(f"₺{total:,.2f}")
        it.setFlags(it.flags() & ~Qt.ItemFlag.ItemIsEditable)
        it.setForeground(QColor("#c62828"))
        self.ppe_table.setItem(row, COL_TOTAL, it)

    def _append_charge_row(self, table, row_ids, db_id, name, qty, unit_price):
        row = table.rowCount()
        table.insertRow(row)
        row_ids.append(db_id)
        table.setItem(row, 0, QTableWidgetItem(str(name)))
        table.setItem(row, 1, QTableWidgetItem(f"{qty:.2f}"))
        table.setItem(row, 2, QTableWidgetItem(f"{unit_price:.2f}"))
        total = qty * unit_price
        it = QTableWidgetItem(f"${total:,.2f}")
        it.setFlags(it.flags() & ~Qt.ItemFlag.ItemIsEditable)
        it.setForeground(QColor("#c62828"))
        table.setItem(row, 3, it)

    def _recalc_ppe(self):
        total = 0.0
        self.ppe_table.blockSignals(True)
        for row in range(self.ppe_table.rowCount()):
            qty   = self._parse_float(self._cell_text(self.ppe_table, row, COL_QTY))
            price = self._parse_float(self._cell_text(self.ppe_table, row, COL_PRICE))
            val   = qty * price
            total += val
            it = QTableWidgetItem(f"₺{val:,.2f}")
            it.setFlags(it.flags() & ~Qt.ItemFlag.ItemIsEditable)
            it.setForeground(QColor("#c62828"))
            self.ppe_table.setItem(row, COL_TOTAL, it)
        self.ppe_table.blockSignals(False)
        self.ppe_total_lbl.setText(f"Total: ₺{total:,.2f}")

    def _recalc_table(self, table: QTableWidget, total_lbl: QLabel):
        total = 0.0
        table.blockSignals(True)
        for row in range(table.rowCount()):
            try:
                qty   = float((table.item(row, 1) or QTableWidgetItem("0")).text())
                price = float((table.item(row, 2) or QTableWidgetItem("0")).text())
            except ValueError:
                qty = price = 0.0
            val = qty * price
            total += val
            it = QTableWidgetItem(f"${val:,.2f}")
            it.setFlags(it.flags() & ~Qt.ItemFlag.ItemIsEditable)
            it.setForeground(QColor("#c62828"))
            table.setItem(row, 3, it)
        table.blockSignals(False)
        total_lbl.setText(f"Total: ${total:,.2f}")

    @staticmethod
    def _cell_text(table, row, col):
        it = table.item(row, col)
        return it.text() if it else ""

    @staticmethod
    def _parse_float(s: str) -> float:
        try:
            return float(s.replace(",", "").replace("₺", "").replace("$", "").strip())
        except ValueError:
            return 0.0

    # ── PPE Import from Excel file ─────────────────────────────────────────────

    def _import_ppe_excel(self):
        if not self.contractor_combo.currentData():
            QMessageBox.warning(self, "No Contractor", "Please select a contractor first.")
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "Open PPE Excel File", "",
            "Excel Files (*.xlsx *.xls *.xlsm);;All Files (*)"
        )
        if not path:
            return
        try:
            rows = self._read_excel_rows(path)
            self._load_excel_rows_into_table(rows)
        except Exception as e:
            QMessageBox.critical(self, "Import Error", str(e))

    def _read_excel_rows(self, path: str) -> list[dict]:
        """Read Excel file and return list of dicts keyed by app column index."""
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)
        # First row = headers
        headers = [str(h or "").strip() for h in next(rows_iter)]
        col_map = []  # position → app column index (or None to skip)
        for h in headers:
            col_map.append(_match_excel_col(h))

        result = []
        for raw_row in rows_iter:
            if all(v is None or str(v).strip() == "" for v in raw_row):
                continue  # skip blank rows
            entry = {}
            for pos, val in enumerate(raw_row):
                if pos >= len(col_map):
                    break
                app_col = col_map[pos]
                if app_col is None or app_col == -1:
                    continue
                entry[app_col] = val
            result.append(entry)
        return result

    def _load_excel_rows_into_table(self, rows: list[dict]):
        self.ppe_table.blockSignals(True)
        for entry in rows:
            row_idx = self.ppe_table.rowCount()
            self.ppe_table.insertRow(row_idx)
            self._ppe_row_ids.append(None)
            for col, val in entry.items():
                self.ppe_table.setItem(row_idx, col, QTableWidgetItem(str(val) if val is not None else ""))
            # init total cell
            it = QTableWidgetItem("₺0.00")
            it.setFlags(it.flags() & ~Qt.ItemFlag.ItemIsEditable)
            it.setForeground(QColor("#c62828"))
            self.ppe_table.setItem(row_idx, COL_TOTAL, it)
        self.ppe_table.blockSignals(False)
        self._recalc_ppe()

    # ── PPE Paste from clipboard (copy from Excel) ─────────────────────────────

    def _paste_ppe_clipboard(self):
        if not self.contractor_combo.currentData():
            QMessageBox.warning(self, "No Contractor", "Please select a contractor first.")
            return
        text = QApplication.clipboard().text()
        if not text.strip():
            QMessageBox.information(self, "Empty Clipboard",
                                    "Nothing found in clipboard. Copy rows from Excel first.")
            return

        lines = text.strip().split("\n")
        if not lines:
            return

        # Detect if first line is a header row (contains known header keywords)
        first = lines[0].lower()
        has_header = any(k in first for k in ("malzeme", "miktar", "birim", "fiyat", "tutar", "kayıt", "kayit"))
        if has_header:
            header_cells = lines[0].split("\t")
            col_map = [_match_excel_col(h) for h in header_cells]
            data_lines = lines[1:]
        else:
            # No header: assume column order matches Excel layout
            # Malzeme | Malzeme kısa metni | Miktar | Giriş ölçü birimi | Kayıt tarihi | Birim Fiyat | Tutar
            col_map = [COL_MAT_CODE, COL_ITEM, COL_QTY, COL_UNIT, COL_DATE, COL_PRICE, None]
            data_lines = lines

        self.ppe_table.blockSignals(True)
        added = 0
        for line in data_lines:
            cells = line.rstrip("\r").split("\t")
            if all(c.strip() == "" for c in cells):
                continue
            row_idx = self.ppe_table.rowCount()
            self.ppe_table.insertRow(row_idx)
            self._ppe_row_ids.append(None)
            for pos, cell_val in enumerate(cells):
                if pos >= len(col_map):
                    break
                app_col = col_map[pos]
                if app_col is None or app_col == -1:
                    continue
                self.ppe_table.setItem(row_idx, app_col, QTableWidgetItem(cell_val.strip()))
            it = QTableWidgetItem("₺0.00")
            it.setFlags(it.flags() & ~Qt.ItemFlag.ItemIsEditable)
            it.setForeground(QColor("#c62828"))
            self.ppe_table.setItem(row_idx, COL_TOTAL, it)
            added += 1
        self.ppe_table.blockSignals(False)
        self._recalc_ppe()
        QMessageBox.information(self, "Pasted", f"{added} row(s) added. Click 'Save All' to store them.")

    # ── PPE helpers ───────────────────────────────────────────────────────────

    def _add_ppe_row(self):
        if not self.contractor_combo.currentData():
            return
        self.ppe_table.blockSignals(True)
        self._append_ppe_row(None, "", "", 0, "", "", 0)
        self.ppe_table.blockSignals(False)

    def _delete_ppe_row(self):
        row = self.ppe_table.currentRow()
        if row < 0:
            return
        db_id = self._ppe_row_ids[row]
        if db_id:
            reply = QMessageBox.question(self, "Confirm", "Delete this row?",
                                         QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply != QMessageBox.StandardButton.Yes:
                return
            m.delete_ppe_charge(db_id)
        self.ppe_table.removeRow(row)
        self._ppe_row_ids.pop(row)
        self._recalc_ppe()

    def _save_ppe(self):
        cid = self.contractor_combo.currentData()
        month = self.month_combo.currentData()
        year = self.year_combo.currentData()
        if not cid:
            return
        errors = []
        for row in range(self.ppe_table.rowCount()):
            item_name = self._cell_text(self.ppe_table, row, COL_ITEM).strip()
            if not item_name:
                errors.append(f"Row {row + 1}: Description is empty.")
                continue
            qty   = self._parse_float(self._cell_text(self.ppe_table, row, COL_QTY))
            price = self._parse_float(self._cell_text(self.ppe_table, row, COL_PRICE))
            mat   = self._cell_text(self.ppe_table, row, COL_MAT_CODE).strip()
            unit  = self._cell_text(self.ppe_table, row, COL_UNIT).strip()
            edate = self._cell_text(self.ppe_table, row, COL_DATE).strip()
            db_id = self._ppe_row_ids[row]
            new_id = m.upsert_ppe_charge(db_id, cid, month, year,
                                         item_name, qty, price,
                                         material_code=mat,
                                         unit_of_measure=unit,
                                         entry_date=edate)
            self._ppe_row_ids[row] = new_id
        if errors:
            QMessageBox.warning(self, "Validation", "\n".join(errors))
        else:
            QMessageBox.information(self, "Saved", "PPE charges saved.")
        self._load_charges()

    # ── Diesel helpers ────────────────────────────────────────────────────────

    def _add_diesel_row(self):
        if not self.contractor_combo.currentData():
            return
        self.diesel_table.blockSignals(True)
        self._append_charge_row(self.diesel_table, self._diesel_row_ids, None, "", 0, 0)
        self.diesel_table.blockSignals(False)

    def _delete_diesel_row(self):
        row = self.diesel_table.currentRow()
        if row < 0:
            return
        db_id = self._diesel_row_ids[row]
        if db_id:
            reply = QMessageBox.question(self, "Confirm", "Delete this row?",
                                         QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply != QMessageBox.StandardButton.Yes:
                return
            m.delete_diesel_charge(db_id)
        self.diesel_table.removeRow(row)
        self._diesel_row_ids.pop(row)
        self._recalc_table(self.diesel_table, self.diesel_total_lbl)

    def _save_diesel(self):
        cid = self.contractor_combo.currentData()
        month = self.month_combo.currentData()
        year = self.year_combo.currentData()
        if not cid:
            return
        errors = []
        for row in range(self.diesel_table.rowCount()):
            desc = (self.diesel_table.item(row, 0) or QTableWidgetItem("")).text().strip()
            if not desc:
                errors.append(f"Row {row + 1}: Description is empty.")
                continue
            try:
                qty   = float((self.diesel_table.item(row, 1) or QTableWidgetItem("0")).text())
                price = float((self.diesel_table.item(row, 2) or QTableWidgetItem("0")).text())
            except ValueError:
                errors.append(f"Row {row + 1}: Invalid number.")
                continue
            db_id = self._diesel_row_ids[row]
            new_id = m.upsert_diesel_charge(db_id, cid, month, year, desc, qty, price)
            self._diesel_row_ids[row] = new_id
        if errors:
            QMessageBox.warning(self, "Validation", "\n".join(errors))
        else:
            QMessageBox.information(self, "Saved", "Diesel charges saved.")
        self._load_charges()
